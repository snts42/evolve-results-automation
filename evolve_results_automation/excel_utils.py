import os
import re
import logging
from collections import Counter, defaultdict
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import COLUMNS, ANALYTICS_FILE, get_excel_file_for_year, list_year_excel_files
from .parsing_utils import unique_row_hash


def _normalize(val):
    """Normalize a cell value to a clean string (empty string for None/NaN)."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def initialize_excel(filepath: str):
    if not os.path.exists(filepath):
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        logging.info(f"Excel file not found, creating {filepath}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(COLUMNS)
        wb.save(filepath)
        wb.close()


def load_existing_results(filepath: str):
    """Load rows from an Excel file as a list of dicts with string values."""
    if not os.path.exists(filepath):
        return []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # First row is the header
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []

        result = []
        for row_values in rows_iter:
            row_dict = {h: _normalize(v) for h, v in zip(headers, row_values)}
            result.append(row_dict)

        return result
    finally:
        wb.close()


def _build_results_workbook(rows: list):
    """Build an in-memory Workbook with the Results sheet. Returns the unsaved wb."""
    date_cols_ddmmyyyy = ["Result Sent", "Certificate", "E-Certificate sent"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.sheet_properties.tabColor = 'E30613'
    ws.append(COLUMNS)
    for row in rows:
        values = []
        for col in COLUMNS:
            val = row.get(col, "")
            if col in date_cols_ddmmyyyy:
                val = format_ddmmyyyy(val)
            values.append(val)
        ws.append(values)

    # Apply formatting in-memory before saving (single I/O operation)
    ws.auto_filter.ref = ws.dimensions
    # Freeze top row so header stays visible when scrolling
    ws.freeze_panes = "A2"
    # Style header row (City & Guilds red with white bold text)
    for cell in ws[1]:
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
    # Alternating row stripes + thin grey borders (consistent with Analytics tab)
    for row_idx in range(2, ws.max_row + 1):
        for cell in ws[row_idx]:
            cell.border = _THIN_BORDER
            cell.font = _DATA_FONT
            if row_idx % 2 == 0:
                cell.fill = _STRIPE_FILL

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 50))

    return wb


def format_ddmmyyyy(val):
    if not val:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    try:
        if len(s) == 10 and s[2] == '/' and s[5] == '/':
            return s
        dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return s


def save_year_to_excel(year, rows_by_year, silent=False):
    """Save rows for a given year to Excel, merging with existing data and deduplicating."""
    if year not in rows_by_year:
        return
    year_rows = rows_by_year[year]
    excel_file = get_excel_file_for_year(year)
    initialize_excel(excel_file)

    # Load existing data (may contain rows from previous accounts)
    existing = load_existing_results(excel_file)

    # Merge existing + new
    combined = existing + year_rows

    # Remove garbage rows missing core fields
    core_fields = ["Completed", "First name", "Last name"]
    combined = [r for r in combined
                if all(r.get(cf, "").strip() for cf in core_fields)]

    # Deduplicate using unique_row_hash, keep last (in-memory rows have latest PDF links)
    seen = {}
    for row in combined:
        seen[unique_row_hash(row)] = row
    combined = list(seen.values())

    # Enforce column order from config (fill missing columns with empty string)
    combined = [{col: row.get(col, "") for col in COLUMNS} for row in combined]

    # Sort by completion date
    def sort_key(row):
        try:
            return datetime.strptime(row.get("Completed", ""), "%d/%m/%Y")
        except (ValueError, TypeError):
            return datetime.max
    try:
        combined.sort(key=sort_key)
    except Exception as e:
        logging.warning(f"Sorting failed for year {year}: {e}")

    wb = _build_results_workbook(combined)
    try:
        wb.save(excel_file)
    finally:
        wb.close()
    if not silent:
        logging.info(f"Saved {len(combined)} rows to {year}/exam_results_{year}.xlsx")


def load_all_existing_data(silent=False):
    """Load all existing year Excel files and return hashes + rows needing PDF download.
    
    Returns:
        tuple: (existing_hashes: set, rows_by_year: dict, pdf_resume_count: int)
    """
    existing_hashes = set()
    rows_by_year = {}
    pdf_resume_count = 0
    for _year_str, excel_path in list_year_excel_files():
        rows = load_existing_results(excel_path)
        for r in rows:
            # Skip garbage rows missing core fields
            completed_val = r.get("Completed", "").strip()
            if not completed_val:
                continue

            existing_hashes.add(unique_row_hash(r))
            # Only add rows needing PDF download to rows_by_year
            report_dl = r.get("PDF report save time", "").strip()
            if not report_dl:
                try:
                    yr = datetime.strptime(completed_val, "%d/%m/%Y").year
                except (ValueError, TypeError):
                    continue  # Skip rows with unparseable dates
                if yr not in rows_by_year:
                    rows_by_year[yr] = []
                rows_by_year[yr].append(r)
                pdf_resume_count += 1
    if not silent:
        logging.info(f"Loaded {len(existing_hashes)} existing results from Excel files")
    return existing_hashes, rows_by_year, pdf_resume_count


# =========================================================================
# ANALYTICS SHEET - Compact dashboard, side-by-side charts
# =========================================================================

# ── Colour palette ────────────────────────────────────────────────────────
_RED      = 'E30613'
_WHITE    = 'FFFFFF'
_BLACK    = '1A1A1A'
_GREY_600 = '757575'
_GREY_400 = 'BDBDBD'
_RED_BG   = 'FFEBEE'
_RED_FG   = 'C62828'

# ── Fonts / fills ─────────────────────────────────────────────────────────
_TITLE_FONT      = Font(bold=True, size=14, color=_WHITE)
_TITLE_META_FONT = Font(size=10, color=_WHITE)
_SECTION_FONT    = Font(bold=True, size=12, color=_BLACK)
_HDR_FONT        = Font(bold=True, size=11, color=_WHITE)
_HDR_FILL        = PatternFill('solid', fgColor=_RED)
_DATA_FONT       = Font(size=11, color=_BLACK)
_STRIPE_FILL     = PatternFill('solid', fgColor='FFF2F2')
_TITLE_FILL      = PatternFill('solid', fgColor=_RED)
_KPI_LABEL_FONT  = Font(bold=True, size=10, color=_GREY_600)
_KPI_VALUE_FONT  = Font(bold=True, size=14, color=_RED)
_KPI_VALUE_SM    = Font(bold=True, size=12, color=_RED)
_KPI_SUB_FONT    = Font(size=8, color=_GREY_600)
_KPI_FILL        = PatternFill('solid', fgColor='FFF2F2')
_INSIGHT_FONT    = Font(size=11, color=_BLACK)
_FOOTER_FONT     = Font(italic=True, size=7, color=_GREY_600)
_THIN_BORDER = Border(
    left=Side('thin', color=_GREY_400), right=Side('thin', color=_GREY_400),
    top=Side('thin', color=_GREY_400), bottom=Side('thin', color=_GREY_400))

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_MONTHS_FULL = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]

# ── Column layout (12 cols: A-L) ─────────────────────────────────────────
_NUM_COLS = 12
_COL_W = {1: 20}
for _ci in range(2, _NUM_COLS + 1):
    _COL_W[_ci] = 13

# Chart sizing
_CHART_W = 14.5          # cm per chart (side-by-side)
_CHART_H = 7.5           # cm height
_CHART_ROWS = 16         # rows a 7.5cm chart occupies

# Wide table column layouts: list of (start_col, end_col) per field
_RESIT_LAYOUT  = [(1, 3), (4, 5), (6, 9), (10, 10), (11, 12)]
_REBOOK_LAYOUT = [(1, 2), (3, 4), (5, 8), (9, 9), (10, 10), (11, 12)]
_EXTRA_LAYOUT  = [(1, 3), (4, 5), (6, 10), (11, 12)]


def _apply_col_widths(ws):
    for ci in range(1, _NUM_COLS + 1):
        ws.column_dimensions[get_column_letter(ci)].width = _COL_W.get(ci, 13)


def _date_key(r):
    """Parse Completed date from row dict, return datetime (or datetime.min if invalid)."""
    try:
        return datetime.strptime(r.get("Completed", ""), "%d/%m/%Y")
    except (ValueError, TypeError):
        return datetime.min


def _group_by_candidate_exam(rows, include_name=False):
    """Group rows by (Enrolment no., Test Name) or (Enrolment no., First name, Last name, Test Name).
    Returns dict mapping key tuple to list of row dicts."""
    pairs = defaultdict(list)
    for rd in rows:
        enrol = rd.get("Enrolment no.", "").strip()
        test = rd.get("Test Name", "").strip()
        if not enrol or not test:
            continue
        if include_name:
            fname = rd.get("First name", "").strip()
            lname = rd.get("Last name", "").strip()
            pairs[(enrol, fname, lname, test)].append(rd)
        else:
            pairs[(enrol, test)].append(rd)
    return pairs


def _compute_monthly(rows):
    """Compute monthly distribution from rows. Returns dict {month_int: {"count": int}}."""
    monthly = {}
    for rd in rows:
        try:
            mi = datetime.strptime(rd.get("Completed", ""), "%d/%m/%Y").month
            if mi not in monthly:
                monthly[mi] = {"count": 0}
            monthly[mi]["count"] += 1
        except (ValueError, TypeError):
            pass
    return monthly


def _compute_centres(rows):
    """Group rows by centre name. Returns defaultdict {centre_name: [rows]}."""
    centres = defaultdict(list)
    for rd in rows:
        cn = rd.get("Centre Name", "").strip() or "Unknown Centre"
        centres[cn].append(rd)
    return centres


def _exam_short(name):
    """Strip numeric prefix from exam name."""
    return re.sub(r'^\d{4}-\d{3}\s*', '', name)


def _exam_chart_label(name, max_line=18):
    """Shorten exam name for chart axis and force 2-line wrap if long."""
    s = _exam_short(name)
    s = s.replace('Functional Skills ', 'FS ')
    if len(s) <= max_line:
        return s
    # Insert newline at best word boundary near midpoint
    mid = len(s) // 2
    best = s.rfind(' ', 0, mid + 5)
    if best <= 0:
        best = s.find(' ', mid)
    if best > 0:
        return s[:best] + '\n' + s[best + 1:]
    return s


# ── Section / table helpers ───────────────────────────────────────────────

def _section(ws, row, title, start_col=1, end_col=_NUM_COLS):
    """Write a section title row, merged across given columns."""
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    ws.cell(row=row, column=start_col, value=title).font = _SECTION_FONT
    ws.row_dimensions[row].height = 22
    return row + 1


def _wide_tbl_header(ws, row, headers, layout):
    """Write header row with merged cells spanning A-L."""
    for (sc, ec), text in zip(layout, headers):
        if ec > sc:
            ws.merge_cells(start_row=row, start_column=sc,
                           end_row=row, end_column=ec)
        c = ws.cell(row=row, column=sc, value=text)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.border = _THIN_BORDER
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        for ci in range(sc + 1, ec + 1):
            cell = ws.cell(row=row, column=ci)
            cell.fill = _HDR_FILL
            cell.border = _THIN_BORDER
    ws.row_dimensions[row].height = 22


def _wide_tbl_row(ws, row, values, layout, stripe=False):
    """Write data row with merged cells spanning A-L."""
    for idx, ((sc, ec), val) in enumerate(zip(layout, values)):
        if ec > sc:
            ws.merge_cells(start_row=row, start_column=sc,
                           end_row=row, end_column=ec)
        c = ws.cell(row=row, column=sc, value=val)
        c.font = _DATA_FONT
        c.border = _THIN_BORDER
        c.alignment = Alignment(
            horizontal='left' if idx == 0 else 'center',
            vertical='center', wrap_text=True)
        if stripe:
            c.fill = _STRIPE_FILL
        for ci in range(sc + 1, ec + 1):
            cell = ws.cell(row=row, column=ci)
            cell.border = _THIN_BORDER
            if stripe:
                cell.fill = _STRIPE_FILL
    ws.row_dimensions[row].height = 20


def _short_centre_name(full_name):
    """Shorten centre name for display (max 31 chars)."""
    name = full_name.strip()
    name = re.sub(r'^\d+\s*\([^)]*\)\s*', '', name)
    if ' - ' in name:
        name = name.split(' - ')[0].strip()
    if not name:
        name = full_name.strip()
    return name[:31]


# ── Chart builder ─────────────────────────────────────────────────────────

def _bar_chart(data_ref, cat_ref, fill_hex=_RED, width=15, height=7.5,
               data_labels=True, y_max=None):
    """Create a styled BarChart object."""
    ch = BarChart()
    ch.type = "col"
    ch.style = 10
    ch.title = None
    ch.y_axis.title = None
    ch.y_axis.scaling.min = 0
    if y_max is not None:
        ch.y_axis.scaling.max = y_max
    ch.y_axis.numFmt = '0'
    ch.y_axis.delete = False
    ch.x_axis.tickLblPos = 'low'
    ch.x_axis.delete = False
    ch.width = width
    ch.height = height
    ch.legend = None
    ch.gapWidth = 80
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cat_ref)
    s = ch.series[0]
    s.graphicalProperties.solidFill = fill_hex
    if data_labels:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.showSerName = False
        s.dLbls.showCatName = False
        s.dLbls.showPercent = False
        s.dLbls.showLegendKey = False
        s.dLbls.numFmt = '0'
    return ch


def _exam_breakdown_chart(dws, exam_d_start, exam_d_end, width=15, height=7.5):
    """Build horizontal bar chart: exam names on Y-axis, counts as bars."""
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont

    ch = BarChart()
    ch.type = "bar"          # horizontal bars
    ch.style = 10
    ch.title = None
    ch.x_axis.title = None   # value axis (bottom)
    ch.x_axis.scaling.min = 0
    ch.x_axis.numFmt = '0'
    ch.x_axis.delete = False
    ch.y_axis.title = None    # category axis (left - exam names)
    ch.y_axis.tickLblPos = 'low'
    ch.y_axis.delete = False
    ch.width = width
    ch.height = height
    ch.legend = None
    ch.gapWidth = 60

    ch.y_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(latin=DrawingFont(typeface='Calibri'), sz=700)
        ), endParaRPr=CharacterProperties(latin=DrawingFont(typeface='Calibri'), sz=700))]
    )

    data_ref = Reference(dws, min_col=2, min_row=exam_d_start,
                         max_row=exam_d_end)
    cat_ref = Reference(dws, min_col=1, min_row=exam_d_start + 1,
                        max_row=exam_d_end)
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cat_ref)

    s = ch.series[0]
    s.graphicalProperties.solidFill = _RED
    s.dLbls = DataLabelList()
    s.dLbls.showVal = True
    s.dLbls.showSerName = False
    s.dLbls.showCatName = False
    s.dLbls.showPercent = False
    s.dLbls.showLegendKey = False
    s.dLbls.numFmt = '0'
    return ch


# ── Compute helpers ───────────────────────────────────────────────────────


def _compute_extra_time(rows):
    """Return list of dicts for candidates whose duration exceeds mode."""
    by_test = defaultdict(list)
    for rd in rows:
        dur_str = rd.get("Duration", "").strip()
        test = rd.get("Test Name", "").strip()
        if dur_str and test:
            try:
                dur_val = float(dur_str)
                by_test[test].append((rd, dur_val))
            except (ValueError, TypeError):
                pass

    result = []
    for test, entries in by_test.items():
        durations = [d for _, d in entries]
        mode_dur = Counter(durations).most_common(1)[0][0] if durations else 0
        if not mode_dur:
            continue
        for rd, dur in entries:
            if dur > mode_dur * 1.1:
                name = (f"{rd.get('First name', '')} "
                        f"{rd.get('Last name', '')}").strip()
                enrol = rd.get("Enrolment no.", "").strip()
                result.append({
                    "name": name, "enrolment": enrol, "exam": test,
                    "extra_pct": f"+{round((dur / mode_dur - 1) * 100)}%",
                    "centre": rd.get("Centre Name", "").strip(),
                })
    return result


def _compute_insights(rows, by_exam, monthly, centres, rebook_opps):
    """Generate auto-insight bullet strings."""
    total = len(rows)
    insights = []
    if by_exam:
        top = by_exam[0]
        pct = round(top["count"] / total * 100) if total else 0
        insights.append(
            f'{top["name"]} is the most sat exam - '
            f'{top["count"]} sittings ({pct}% of annual volume)')

    if rebook_opps:
        insights.append(
            f'{len(rebook_opps)} candidate(s) failed and have not rebooked - '
            f'potential revenue to recover')

    busiest = max(monthly.items(), key=lambda x: x[1]["count"],
                  default=(0, {"count": 0}))
    if busiest[1]["count"] > 0:
        mi = busiest[0]
        cnt = busiest[1]["count"]
        insights.append(
            f'{_MONTHS_FULL[mi - 1]} was the busiest month - {cnt} sittings')

    return insights


# ── Header / KPI / footer builders ────────────────────────────────────────

def _build_header(ws, title, subtitle, year):
    """Write single-row red header: bold title left, smaller meta right."""
    r = 1
    # Title in A-D
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value=title)
    c.font = _TITLE_FONT
    c.fill = _TITLE_FILL
    c.alignment = Alignment(vertical='center')
    for ci in range(2, 5):
        ws.cell(row=r, column=ci).fill = _TITLE_FILL
    # Meta in E-L (smaller font)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=_NUM_COLS)
    meta = (f"{subtitle}    Year: {year}    "
            f"Generated: {datetime.now().strftime('%d/%m/%Y')}")
    m = ws.cell(row=r, column=5, value=meta)
    m.font = _TITLE_META_FONT
    m.fill = _TITLE_FILL
    m.alignment = Alignment(vertical='center')
    for ci in range(6, _NUM_COLS + 1):
        ws.cell(row=r, column=ci).fill = _TITLE_FILL
    ws.row_dimensions[r].height = 30
    return r + 2  # row 2 = spacer


def _build_kpis(ws, row, kpis, layout=None):
    """Write KPI cards with grey card background and red top accent.
    layout = list of (start_col, end_col) per KPI. Returns next row."""
    if layout is None:
        layout = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]

    for i, kpi in enumerate(kpis):
        if i >= len(layout):
            break
        sc, ec = layout[i]
        span = ec - sc + 1
        label, val = kpi[0], kpi[1]
        sub = kpi[2] if len(kpi) > 2 else None
        # Use smaller font for text values that need more space
        is_text = isinstance(val, str) and len(val) > 6
        val_font = _KPI_VALUE_SM if is_text else _KPI_VALUE_FONT

        # Merge cells for 3 rows
        for r_off in range(3):
            ws.merge_cells(start_row=row + r_off, start_column=sc,
                           end_row=row + r_off, end_column=ec)

        # Apply card styling: grey fill + red top/bottom accent
        for r_off in range(3):
            for c_off in range(span):
                cell = ws.cell(row=row + r_off, column=sc + c_off)
                cell.fill = _KPI_FILL
                top_s = Side('medium', color=_RED) if r_off == 0 else Side('thin', color=_GREY_400)
                bot_s = Side('medium', color=_RED) if r_off == 2 else Side('thin', color=_GREY_400)
                cell.border = Border(
                    top=top_s, bottom=bot_s,
                    left=Side('thin', color=_GREY_400),
                    right=Side('thin', color=_GREY_400))

        # Label row
        lc = ws.cell(row=row, column=sc, value=label)
        lc.font = _KPI_LABEL_FONT
        lc.fill = _KPI_FILL
        lc.alignment = Alignment(horizontal='center', vertical='center')

        # Value row
        vc = ws.cell(row=row + 1, column=sc, value=val)
        vc.font = val_font
        vc.fill = _KPI_FILL
        vc.alignment = Alignment(horizontal='center', vertical='center',
                                      wrap_text=True)

        # Subtitle row
        if sub:
            sc2 = ws.cell(row=row + 2, column=sc, value=sub)
            sc2.font = _KPI_SUB_FONT
            sc2.fill = _KPI_FILL
            sc2.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)

    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 38
    ws.row_dimensions[row + 2].height = 20
    return row + 4


def _build_footer(ws, row, text):
    """Write a grey footer line."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NUM_COLS)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _FOOTER_FONT
    c.alignment = Alignment(horizontal='left')
    ws.row_dimensions[row].height = 14
    return row + 1


# ── Public entry-point ────────────────────────────────────────────────────
def add_analytics_sheet(filepath, rows=None, all_rows=None):
    """Add/replace the Analytics sheet on an existing Excel file.
    If all_rows is provided, rebook opportunities cross-reference all years."""
    if rows is None:
        rows = load_existing_results(filepath)
    if not rows:
        return
    wb = load_workbook(filepath)
    try:
        if wb.active and wb.active.title not in ("Results", "Analytics"):
            wb.active.title = "Results"
        _add_analytics_to_wb(wb, rows, all_rows=all_rows)
        wb.save(filepath)
    finally:
        wb.close()


# ── Shared year-level dashboard builder ───────────────────────────────────
def _build_year_dashboard(wb, ws, rows, all_rows, year,
                          chart_sheet_name, header_title, footer_text):
    """Build a full year analytics dashboard on the given worksheet.
    Used by both per-year Excel analytics and combined workbook year tabs."""
    centres = _compute_centres(rows)
    total = len(rows)
    by_exam = _compute_by_exam(rows)
    extra_time = _compute_extra_time(rows)
    monthly = _compute_monthly(rows)

    rebook_source = all_rows if all_rows else rows
    rebook_opps = _compute_rebook_opportunities(rebook_source, year_filter=year)

    customers = len({rd.get("Enrolment no.", "").strip() for rd in rows
                     if rd.get("Enrolment no.", "").strip()})

    # Resit conversion (cross-year aware)
    resits_returned, unique_fails = _compute_resit_conversion(rebook_source, year)
    resit_conv = (f"{round(resits_returned / unique_fails * 100)}%"
                  if unique_fails else "N/A")
    resit_sub = (f"{resits_returned} of {unique_fails} returned"
                 if unique_fails else "No fails")

    rebook_count = len(rebook_opps)
    rebook_sub = (f"{rebook_count} of {unique_fails} fails outstanding"
                  if unique_fails else "No fails")

    top_exam = _exam_short(by_exam[0]["name"]) if by_exam else "N/A"
    top_exam_pct = round(by_exam[0]["count"] / total * 100) if by_exam and total else 0
    top_exam_sub = f"{by_exam[0]['count']} sittings ({top_exam_pct}%)" if by_exam else ""

    busiest_mi = max(range(1, 13),
                     key=lambda m: monthly.get(m, {"count": 0})["count"],
                     default=1)
    busiest_cnt = monthly.get(busiest_mi, {"count": 0})["count"]
    busiest_pct = round(busiest_cnt / total * 100) if total else 0
    busiest_sub = f"{busiest_cnt} sittings ({busiest_pct}%)" if busiest_cnt else ""

    # ── Hidden chart-data sheet ───────────────────────────────────────
    dws = wb.create_sheet(chart_sheet_name)
    dws.sheet_state = 'hidden'

    dws.cell(1, 1, "Month")
    dws.cell(1, 2, "Exams")
    for i in range(12):
        dws.cell(2 + i, 1, _MONTHS[i])
        cnt = monthly.get(i + 1, {"count": 0})["count"]
        dws.cell(2 + i, 2, cnt if cnt else None)

    exam_d_start = 15
    dws.cell(exam_d_start, 1, "Exam")
    dws.cell(exam_d_start, 2, "Count")
    for i, ex in enumerate(by_exam):
        dws.cell(exam_d_start + 1 + i, 1, _exam_chart_label(ex["name"]))
        dws.cell(exam_d_start + 1 + i, 2, ex["count"])
    exam_d_end = exam_d_start + len(by_exam)

    # ── Header + KPIs ─────────────────────────────────────────────────
    r = _build_header(ws, header_title,
                      f" All Centres ({len(centres)})", year)

    r = _section(ws, r, "OVERVIEW")
    kpi_layout = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]
    r = _build_kpis(ws, r, [
        ("Total Exams", total),
        ("Unique Candidates", customers),
        ("Resit Conversion", resit_conv, resit_sub),
        ("Most Popular Exam", top_exam, top_exam_sub),
        ("Busiest Month", _MONTHS_FULL[busiest_mi - 1], busiest_sub),
        ("Rebook Opportunities", rebook_count, rebook_sub),
    ], layout=kpi_layout)

    # ── Side-by-side charts ───────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="MONTHLY VOLUME").font = _SECTION_FONT
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=12)
    ws.cell(row=r, column=7, value="EXAM BREAKDOWN").font = _SECTION_FONT
    ws.row_dimensions[r].height = 22
    r += 1

    chart_anchor_row = r
    month_max = max((monthly.get(m, {"count": 0})["count"]
                     for m in range(1, 13)), default=0)
    ch_month = _bar_chart(
        Reference(dws, min_col=2, min_row=1, max_row=13),
        Reference(dws, min_col=1, min_row=2, max_row=13),
        fill_hex=_RED, width=_CHART_W, height=_CHART_H,
        y_max=int(month_max * 1.25) + 1 if month_max else None)
    ws.add_chart(ch_month, f"A{chart_anchor_row}")

    if by_exam:
        ch_exam = _exam_breakdown_chart(
            dws, exam_d_start, exam_d_end,
            width=_CHART_W, height=_CHART_H)
        ws.add_chart(ch_exam, f"G{chart_anchor_row}")

    r = chart_anchor_row + _CHART_ROWS
    r += 1  # spacer

    # ── Key Insights ──────────────────────────────────────────────────
    insights = _compute_insights(rows, by_exam, monthly, centres, rebook_opps)
    if insights:
        r = _section(ws, r, "KEY INSIGHTS")
        for ins in insights:
            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r, end_column=_NUM_COLS)
            c = ws.cell(row=r, column=1, value=f"  {ins}")
            c.font = _INSIGHT_FONT
            c.alignment = Alignment(vertical='center', wrap_text=True)
            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

    # ── Rebook Opportunities (summary by exam) ────────────────────────
    r = _section(ws, r, "REBOOK OPPORTUNITIES")
    if rebook_opps:
        exam_summary = defaultdict(lambda: {"count": 0, "total_days": 0})
        for rb in rebook_opps:
            es = exam_summary[rb["exam"]]
            es["count"] += 1
            es["total_days"] += rb["days_ago"]
        summary_layout = [(1, 4), (5, 7), (8, 10), (11, 12)]
        _wide_tbl_header(ws, r, ["Exam", "Candidates", "Avg Days Ago",
                                  "% of Total"], summary_layout)
        r += 1
        for idx, (exam, s) in enumerate(sorted(exam_summary.items(),
                                                key=lambda x: x[1]["count"],
                                                reverse=True)):
            avg_days = round(s["total_days"] / s["count"]) if s["count"] else 0
            pct = f"{round(s['count'] / len(rebook_opps) * 100)}%"
            _wide_tbl_row(ws, r, [_exam_short(exam), s["count"],
                                   avg_days, pct],
                          summary_layout, stripe=(idx % 2 == 1))
            r += 1
        r += 1

        # ── Candidates to Rebook (detail table) ──────────────────────
        r = _section(ws, r, "CANDIDATES TO REBOOK")
        _wide_tbl_header(ws, r, ["Candidate", "Enrolment", "Exam",
                                  "Attempts", "Last Failed", "Days Ago"],
                         _REBOOK_LAYOUT)
        r += 1
        for idx, rb in enumerate(rebook_opps):
            _wide_tbl_row(ws, r, [rb["name"], rb["enrolment"],
                                   _exam_short(rb["exam"]),
                                   rb["attempts"],
                                   rb["last_fail"], rb["days_ago"]],
                          _REBOOK_LAYOUT, stripe=(idx % 2 == 1))
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=_NUM_COLS)
        ws.cell(row=r, column=1,
                value="  No rebook opportunities - all fails have "
                      "rebooked or passed on resit").font = _INSIGHT_FONT
        r += 1
    r += 1

    # ── Extra Time ────────────────────────────────────────────────────
    r = _section(ws, r, "EXTRA TIME CANDIDATES")
    if extra_time:
        _wide_tbl_header(ws, r, ["Candidate", "Enrolment", "Exam",
                                  "Extra Time"],
                         _EXTRA_LAYOUT)
        r += 1
        for idx, et in enumerate(extra_time):
            _wide_tbl_row(ws, r, [et["name"], et["enrolment"], et["exam"],
                                   et["extra_pct"]],
                          _EXTRA_LAYOUT, stripe=(idx % 2 == 1))
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=_NUM_COLS)
        ws.cell(row=r, column=1,
                value="  No extra time candidates detected").font = _INSIGHT_FONT
        r += 1
    r += 1

    # ── Footer ────────────────────────────────────────────────────────
    _build_footer(ws, r, footer_text)
    _apply_col_widths(ws)


# ── Per-year Excel orchestrator ───────────────────────────────────────────
def _add_analytics_to_wb(wb, rows, all_rows=None):
    """Build the Analytics tab in a per-year Excel file."""
    if not rows:
        return

    # Remove all non-Results tabs
    for name in list(wb.sheetnames):
        if name != "Results":
            del wb[name]

    # Detect year
    year = datetime.now().year
    for rd in rows:
        try:
            year = datetime.strptime(rd.get("Completed", ""), "%d/%m/%Y").year
            break
        except (ValueError, TypeError):
            pass

    ws = wb.create_sheet("Analytics", 1)
    ws.sheet_properties.tabColor = _RED
    ws.sheet_view.showGridLines = False

    centres = _compute_centres(rows)
    centre_names = ", ".join(_short_centre_name(cn)
                             for cn in sorted(centres.keys()))
    footer = (f"Data: Results tab - {len(rows)} records - "
              f"{len(centres)} centre(s): {centre_names} - "
              f"Internal use - City & Guilds")

    _build_year_dashboard(wb, ws, rows, all_rows, year,
                          "_ChartData", "EVOLVE SECURE-ASSESS ANALYTICS",
                          footer)


def _compute_by_exam(rows):
    """Group rows by Test Name. Returns list of dicts sorted by count desc.
    Each: {name, count, passed, failed, rate, avg_score}."""
    groups = defaultdict(list)
    for r in rows:
        name = r.get("Test Name", "").strip()
        if name:
            groups[name].append(r)
    result = []
    for name, group in groups.items():
        count = len(group)
        passed = sum(1 for r in group if r.get("Result", "").strip().lower() == "pass")
        scores = []
        for r in group:
            try:
                scores.append(float(r.get("Percent", "0").replace("%", "")))
            except (ValueError, TypeError):
                pass
        avg_score = round(sum(scores) / len(scores), 1) if scores else 0.0
        result.append({
            "name": name, "count": count, "passed": passed,
            "failed": count - passed,
            "rate": round(passed / count * 100, 1) if count else 0.0,
            "avg_score": avg_score,
        })
    result.sort(key=lambda x: x["count"], reverse=True)
    return result


# =========================================================================
# COMBINED ANALYTICS WORKBOOK - Cross-year aggregation
# =========================================================================


def _compute_resit_conversion(all_rows, year):
    """Count how many unique candidate+exam combos that failed in `year`
    went on to have at least one later attempt (in any year).
    Returns (returned, total_unique_fails) so caller can compute the rate."""
    pairs = _group_by_candidate_exam(all_rows, include_name=False)

    returned = 0
    total_unique_fails = 0
    for (enrol, test), attempts in pairs.items():
        attempts.sort(key=_date_key)
        # Find fails in the target year
        fails_in_year = [a for a in attempts
                         if _date_key(a).year == year
                         and a.get("Result", "").strip().lower() == "fail"]
        if not fails_in_year:
            continue
        total_unique_fails += 1
        # Check if there's any attempt AFTER the last fail in that year
        last_fail_dt = _date_key(fails_in_year[-1])
        has_later = any(_date_key(a) > last_fail_dt for a in attempts)
        if has_later:
            returned += 1
    return returned, total_unique_fails


def _compute_rebook_opportunities(all_rows, year_filter=None):
    """Find candidates whose latest attempt for any exam is a fail - potential revenue.
    Includes both single-attempt fails and multi-attempt fails (never passed).
    If year_filter is set, only include candidates who had a fail in that year
    and whose latest overall attempt is still a fail (i.e. they haven't rebooked
    or passed in a later year).
    Sorted by days since fail (most overdue at top)."""
    pairs = _group_by_candidate_exam(all_rows, include_name=True)

    rebooks = []
    for (enrol, fname, lname, test), attempts in pairs.items():
        attempts.sort(key=_date_key)
        latest = attempts[-1]
        latest_result = latest.get("Result", "").strip().lower()
        if latest_result != "fail":
            continue
        # If year_filter is set, only include if the latest fail is in that year
        if year_filter is not None:
            if _date_key(latest).year != year_filter:
                continue
        fail_date = latest.get("Completed", "")
        try:
            days_ago = (datetime.now() - datetime.strptime(fail_date, "%d/%m/%Y")).days
        except (ValueError, TypeError):
            days_ago = 0
        rebooks.append({
            "name": f"{fname} {lname}".strip(),
            "enrolment": enrol,
            "exam": test,
            "attempts": len(attempts),
            "last_fail": fail_date,
            "days_ago": days_ago,
        })
    rebooks.sort(key=lambda x: x["days_ago"], reverse=True)
    return rebooks


def _build_analytics_year_tab(wb, year_str, rows, all_rows=None):
    """Build a single year analytics tab in the combined workbook."""
    ws = wb.create_sheet(year_str)
    ws.sheet_properties.tabColor = _RED
    ws.sheet_view.showGridLines = False

    year = int(year_str)
    centres = _compute_centres(rows)
    centre_names = ", ".join(_short_centre_name(cn)
                             for cn in sorted(centres.keys()))
    footer = (f"Data: {len(rows)} records - "
              f"{len(centres)} centre(s): {centre_names}")

    _build_year_dashboard(wb, ws, rows, all_rows, year,
                          f"_{year_str}_ChartData",
                          f"ANALYTICS - {year_str}", footer)


def _build_analytics_overview_tab(wb, all_rows, rows_by_year):
    """Build the Overview tab with cross-year aggregated analytics."""
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = _RED
    ws.sheet_view.showGridLines = False

    years_sorted = sorted(rows_by_year.keys(), reverse=True)
    total = len(all_rows)
    by_exam = _compute_by_exam(all_rows)
    rebook_opps = _compute_rebook_opportunities(all_rows)

    customers = len({rd.get("Enrolment no.", "").strip() for rd in all_rows
                     if rd.get("Enrolment no.", "").strip()})
    
    total_fails = sum(1 for rd in all_rows
                      if rd.get("Result", "").strip().lower() == "fail")

    # Resit conversion: across all years, how many unique candidate+exam
    # combos that failed went on to have at least one later attempt?
    pairs = _group_by_candidate_exam(all_rows, include_name=False)
    total_unique_fails = 0
    resits_returned = 0
    for (enrol, test), attempts in pairs.items():
        attempts.sort(key=_date_key)
        fails = [a for a in attempts
                 if a.get("Result", "").strip().lower() == "fail"]
        if not fails:
            continue
        total_unique_fails += 1
        last_fail_dt = _date_key(fails[-1])
        if any(_date_key(a) > last_fail_dt for a in attempts):
            resits_returned += 1

    resit_conv = (f"{round(resits_returned / total_unique_fails * 100)}%"
                  if total_unique_fails else "N/A")
    resit_sub = (f"{resits_returned} of {total_unique_fails} returned"
                 if total_unique_fails else "No fails")

    top_exam = _exam_short(by_exam[0]["name"]) if by_exam else "N/A"
    top_exam_pct = round(by_exam[0]["count"] / total * 100) if by_exam and total else 0
    top_exam_sub = f"{by_exam[0]['count']} sittings ({top_exam_pct}%)" if by_exam else ""

    # Monthly data (all years combined)
    monthly = _compute_monthly(all_rows)

    busiest_mi = max(range(1, 13),
                     key=lambda m: monthly.get(m, {"count": 0})["count"],
                     default=1)
    busiest_cnt = monthly.get(busiest_mi, {"count": 0})["count"]
    busiest_pct = round(busiest_cnt / total * 100) if total else 0
    busiest_sub = f"{busiest_cnt} sittings ({busiest_pct}%)" if busiest_cnt else ""

    centres = _compute_centres(all_rows)

    # Header
    year_range = (f"{min(years_sorted)}-{max(years_sorted)}"
                  if len(years_sorted) > 1 else str(years_sorted[0]))
    r = _build_header(ws, "COMBINED ANALYTICS",
                      f" All Centres ({len(centres)})", year_range)

    # KPIs - same style as per-year analytics
    r = _section(ws, r, "OVERVIEW")
    total_pass = total - total_fails
    pass_rate = f"{round(total_pass / total * 100)}%" if total else "N/A"
    pass_sub = f"{total_pass} of {total}" if total else ""
    kpi_layout = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]
    r = _build_kpis(ws, r, [
        ("Total Exams", total),
        ("Unique Candidates", customers),
        ("Overall Pass Rate", pass_rate, pass_sub),
        ("Most Popular Exam", top_exam, top_exam_sub),
        ("Busiest Month", _MONTHS_FULL[busiest_mi - 1], busiest_sub),
        ("Resit Conversion", resit_conv, resit_sub),
    ], layout=kpi_layout)

    # ── Hidden chart data ────────────────────────────────────────────
    dws = wb.create_sheet("_OverviewChartData")
    dws.sheet_state = 'hidden'

    # Year-over-year data (col 1-2)
    dws.cell(1, 1, "Year")
    dws.cell(1, 2, "Exams")
    for i, yr in enumerate(sorted(years_sorted)):
        yr_rows = rows_by_year[yr]
        dws.cell(2 + i, 1, str(yr))
        dws.cell(2 + i, 2, len(yr_rows))
    yoy_end = 1 + len(years_sorted)

    # Monthly data (col 4-5)
    dws.cell(1, 4, "Month")
    dws.cell(1, 5, "Exams")
    for i in range(12):
        dws.cell(2 + i, 4, _MONTHS[i])
        cnt = monthly.get(i + 1, {"count": 0})["count"]
        dws.cell(2 + i, 5, cnt if cnt else None)

    # Exam data (row 16+, col 1-2)
    exam_d_start = 16
    dws.cell(exam_d_start, 1, "Exam")
    dws.cell(exam_d_start, 2, "Count")
    for i, ex in enumerate(by_exam):
        dws.cell(exam_d_start + 1 + i, 1, _exam_chart_label(ex["name"]))
        dws.cell(exam_d_start + 1 + i, 2, ex["count"])
    exam_d_end = exam_d_start + len(by_exam)

    # ── Row 1 charts: YoY volume (left) + Monthly volume (right) ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="YEAR-OVER-YEAR VOLUME").font = _SECTION_FONT
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=12)
    ws.cell(row=r, column=7, value="MONTHLY VOLUME (ALL YEARS)").font = _SECTION_FONT
    ws.row_dimensions[r].height = 22
    r += 1

    chart_anchor_row = r
    if len(years_sorted) > 0:
        yoy_max = max(len(rows_by_year[y]) for y in years_sorted)
        ch_yoy = _bar_chart(
            Reference(dws, min_col=2, min_row=1, max_row=yoy_end),
            Reference(dws, min_col=1, min_row=2, max_row=yoy_end),
            fill_hex=_RED, width=_CHART_W, height=_CHART_H,
            y_max=int(yoy_max * 1.25) + 1 if yoy_max else None)
        ws.add_chart(ch_yoy, f"A{chart_anchor_row}")

    month_max = max((monthly.get(m, {"count": 0})["count"]
                     for m in range(1, 13)), default=0)
    ch_month = _bar_chart(
        Reference(dws, min_col=5, min_row=1, max_row=13),
        Reference(dws, min_col=4, min_row=2, max_row=13),
        fill_hex=_RED, width=_CHART_W, height=_CHART_H,
        y_max=int(month_max * 1.25) + 1 if month_max else None)
    ws.add_chart(ch_month, f"G{chart_anchor_row}")

    r = chart_anchor_row + _CHART_ROWS

    # ── Row 2 chart: Exam breakdown (full width, height scales with exam count)
    r = _section(ws, r, "EXAM BREAKDOWN")
    chart2_anchor = r
    exam_count = len(by_exam)
    exam_chart_h = max(_CHART_H, exam_count * 1.2)
    exam_chart_rows = max(_CHART_ROWS, int(exam_chart_h / _CHART_H * _CHART_ROWS))
    if by_exam:
        ch_exam = _exam_breakdown_chart(
            dws, exam_d_start, exam_d_end,
            width=29, height=exam_chart_h)
        ws.add_chart(ch_exam, f"A{chart2_anchor}")
    r = chart2_anchor + exam_chart_rows

    # ── YoY comparison table (no Centres column) ─────────────────
    r = _section(ws, r, "YEAR-OVER-YEAR COMPARISON")
    yoy_layout = [(1, 3), (4, 5), (6, 7), (8, 9), (10, 12)]
    _wide_tbl_header(ws, r, ["Year", "Exams", "Candidates", "Pass Rate",
                              "Avg Score"],
                     yoy_layout)
    r += 1
    for idx, yr in enumerate(sorted(years_sorted, reverse=True)):
        yr_rows = rows_by_year[yr]
        yr_total = len(yr_rows)
        yr_cust = len({rd.get("Enrolment no.", "").strip() for rd in yr_rows
                       if rd.get("Enrolment no.", "").strip()})
        yr_pass = sum(1 for rd in yr_rows
                      if rd.get("Result", "").strip().lower() == "pass")
        yr_rate = f"{round(yr_pass / yr_total * 100)}%" if yr_total else "N/A"
        yr_scores = []
        for rd in yr_rows:
            try:
                yr_scores.append(float(rd.get("Percent", "0").replace("%", "")))
            except (ValueError, TypeError):
                pass
        yr_avg = f"{round(sum(yr_scores) / len(yr_scores), 1)}%" if yr_scores else "N/A"
        _wide_tbl_row(ws, r, [str(yr), yr_total, yr_cust, yr_rate, yr_avg],
                      yoy_layout, stripe=(idx % 2 == 1))
        r += 1
    r += 1

    # ── Exam breakdown table ─────────────────────────────────────
    r = _section(ws, r, "EXAM BREAKDOWN (ALL YEARS)")
    exam_tbl_layout = [(1, 6), (7, 8), (9, 10), (11, 12)]
    _wide_tbl_header(ws, r, ["Exam", "Sittings", "Pass Rate", "Avg Score"],
                     exam_tbl_layout)
    r += 1
    for idx, ex in enumerate(by_exam):
        _wide_tbl_row(ws, r, [
            _exam_short(ex["name"]), ex["count"],
            f"{ex['rate']}%", f"{ex['avg_score']}%"],
            exam_tbl_layout, stripe=(idx % 2 == 1))
        r += 1
    r += 1

    # ── Rebook Opportunities (summary by exam) ────────────────────────
    r = _section(ws, r, "REBOOK OPPORTUNITIES")
    if rebook_opps:
        exam_summary = defaultdict(lambda: {"count": 0, "total_days": 0})
        for rb in rebook_opps:
            es = exam_summary[rb["exam"]]
            es["count"] += 1
            es["total_days"] += rb["days_ago"]
        summary_layout = [(1, 4), (5, 7), (8, 10), (11, 12)]
        _wide_tbl_header(ws, r, ["Exam", "Candidates", "Avg Days Ago",
                                  "% of Total"], summary_layout)
        r += 1
        for idx, (exam, s) in enumerate(sorted(exam_summary.items(),
                                                key=lambda x: x[1]["count"],
                                                reverse=True)):
            avg_days = round(s["total_days"] / s["count"]) if s["count"] else 0
            pct = f"{round(s['count'] / len(rebook_opps) * 100)}%"
            _wide_tbl_row(ws, r, [_exam_short(exam), s["count"],
                                   avg_days, pct],
                          summary_layout, stripe=(idx % 2 == 1))
            r += 1
        r += 1

        # ── Candidates to Rebook (detail table) ──────────────────────
        r = _section(ws, r, "CANDIDATES TO REBOOK")
        _wide_tbl_header(ws, r, ["Candidate", "Enrolment", "Exam",
                                  "Attempts", "Last Failed", "Days Ago"],
                         _REBOOK_LAYOUT)
        r += 1
        for idx, rb in enumerate(rebook_opps):
            _wide_tbl_row(ws, r, [rb["name"], rb["enrolment"],
                                   _exam_short(rb["exam"]),
                                   rb["attempts"],
                                   rb["last_fail"], rb["days_ago"]],
                          _REBOOK_LAYOUT, stripe=(idx % 2 == 1))
            # Highlight days ago column since they're all actionable fails
            for ci in [11, 12]:
                cell = ws.cell(row=r, column=ci)
                cell.fill = PatternFill('solid', fgColor=_RED_BG)
                cell.font = Font(bold=True, size=9, color=_RED_FG)
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=_NUM_COLS)
        ws.cell(row=r, column=1,
                value="  No rebook opportunities - all fails have "
                      "rebooked or passed on resit").font = _INSIGHT_FONT
        r += 1
    r += 1

    # Footer
    centre_names = ", ".join(_short_centre_name(cn) for cn in sorted(centres.keys()))
    _build_footer(ws, r, f"Data: {total} records across {len(years_sorted)} year(s) - "
                          f"{len(centres)} centre(s): {centre_names} - "
                          f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    _apply_col_widths(ws)


def generate_analytics_workbook(all_rows=None, rows_by_year=None):
    """Generate the combined analytics.xlsx with Overview + per-year tabs.
    Only generates when 2+ years of data exist.
    
    Args:
        all_rows: Optional pre-loaded list of all rows (avoids re-reading files)
        rows_by_year: Optional pre-loaded dict {year: [rows]} (avoids re-reading files)
    """
    # If data not provided, load from files
    if all_rows is None or rows_by_year is None:
        year_files = list_year_excel_files()
        if len(year_files) < 2:
            logging.info("Need 2+ years of data for combined analytics, skipping")
            return

        all_rows = []
        rows_by_year = {}
        for year_str, filepath in year_files:
            rows = load_existing_results(filepath)
            if rows:
                yr = int(year_str)
                rows_by_year[yr] = rows
                all_rows.extend(rows)

    if not all_rows:
        logging.info("No data found across year files, skipping analytics")
        return
    
    if len(rows_by_year) < 2:
        logging.info("Need 2+ years of data for combined analytics, skipping")
        return

    logging.debug(f"Generating analytics from {len(all_rows)} rows across "
                  f"{len(rows_by_year)} year(s)...")

    wb = Workbook()
    try:
        # Tab 1: Overview (aggregated)
        _build_analytics_overview_tab(wb, all_rows, rows_by_year)

        # Tab 2+: One per year (most recent first)
        for yr in sorted(rows_by_year.keys(), reverse=True):
            _build_analytics_year_tab(wb, str(yr), rows_by_year[yr], all_rows)

        wb.save(ANALYTICS_FILE)
        logging.debug(f"Analytics saved to {os.path.basename(ANALYTICS_FILE)}")
    finally:
        wb.close()


def regenerate_analytics():
    """Regenerate per-year and combined analytics from all year Excel files.
    Returns True if analytics were generated, False if no data to process."""
    all_rows = []
    by_year = {}
    for year_str, filepath in list_year_excel_files():
        yr = int(year_str)
        yr_rows = load_existing_results(filepath)
        by_year[yr] = yr_rows
        all_rows.extend(yr_rows)
    if not all_rows:
        return False
    for year in by_year:
        try:
            add_analytics_sheet(get_excel_file_for_year(year), all_rows=all_rows)
        except Exception as e:
            logging.error(f"Failed to generate analytics for {year}: {e}")
    try:
        generate_analytics_workbook(all_rows=all_rows, rows_by_year=by_year)
    except Exception as e:
        logging.error(f"Failed to generate combined analytics: {e}")
    return True
